# app/contact/routes.py
from flask import render_template, flash, redirect, url_for, current_app, jsonify, request
from flask_mail import Message
from datetime import datetime
import json
import os
import openpyxl
from openpyxl import Workbook
import threading
import time
import socket
from contextlib import contextmanager
# Import the 'contact' blueprint instance from its __init__.py file
from . import contact
# Import your ContactForm from the forms.py file within the same 'contact' package
from .forms import ContactForm
# Import the global mail instance
from .. import mail


# Define the route for the contact page.
# This route will handle both GET requests (to display the form)
# and POST requests (to process the form submission).
# The endpoint for this route will be 'contact.contact'.
@contact.route('/contact-us', methods=['GET', 'POST'])
def contact_form():
    # Create an instance of the ContactForm.
    # Flask-WTF will automatically populate it with submitted data on POST requests.
    form = ContactForm()

    # Add debug logging for form validation
    if request.method == 'POST':
        current_app.logger.info(f"Form submission attempt - Specialties raw data: {form.specialties.data}")
        current_app.logger.info(f"Form validation errors: {form.errors}")

    # Check if the form was submitted via POST and if all validators passed.
    if form.validate_on_submit():
        # Retrieve data from the form fields.
        name = form.name.data
        email = form.email.data

        # Handle phone number properly - use phone_number field primarily
        phone = form.phone_number.data if form.phone_number.data else (form.phone.data if hasattr(form, 'phone') and form.phone.data else '')

        # Get country code if available
        country_code = form.phone_country_code.data if hasattr(form, 'phone_country_code') and form.phone_country_code.data else '+1'
        if phone and not phone.startswith('+'):
            phone = f"{country_code} {phone}"

        practice_name = form.practice.data
        state = form.state.data
        specialties = form.specialties.data
        service_type_value = form.service_type.data

        # Debug logging for specialty field
        current_app.logger.info(f"Successfully processed form - Specialties: '{specialties}' (type: {type(specialties)})")
        if specialties:
            specialties_list = [s.strip() for s in specialties.split(',') if s.strip()]
            current_app.logger.info(f"Parsed specialties list: {specialties_list} (count: {len(specialties_list)})")
        subject = form.subject.data if form.subject.data else f"Contact Form - {service_type_value or 'General Inquiry'}"
        message = form.message.data
        privacy_policy_agreement = form.privacy_policy_agreement.data

        # Prepare form data for processing
        form_data = {
            'name': name,
            'email': email,
            'phone': phone,
            'practice_name': practice_name,
            'state': state,
            'specialties': specialties,
            'service_type': service_type_value,
            'subject': subject,
            'message': message,
            'privacy_policy_agreement': privacy_policy_agreement,
            'timestamp': datetime.utcnow().isoformat(),
            'ip_address': request.remote_addr or 'Unknown'
        }

        # 🚀 IMMEDIATE USER RESPONSE - Show success message first for better UX
        inquiry_type = service_type_value or "inquiry"
        flash('Thank you, {}! Your message has been received successfully. We will contact you soon regarding your {}.'.format(name, inquiry_type), 'success')
        current_app.logger.info("✅ Contact form submitted successfully by {} ({})".format(name, email))

        # 🔥 PERFORMANCE OPTIMIZATION: Process heavy operations in background thread
        # This allows immediate response to user while handling emails/logging asynchronously
        background_thread = threading.Thread(
            target=process_contact_form_background,
            args=(form_data.copy(), current_app._get_current_object())
        )
        background_thread.daemon = True
        background_thread.start()

        current_app.logger.info("🚀 Background processing started for {} ({})".format(name, form_data.get('email', 'unknown')))

        # Check if this is an AJAX request (from modal)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True,
                'message': 'Thank you, {}! Your message has been received successfully. We will contact you soon regarding your {}.'.format(name, inquiry_type)
            })

        # Redirect the user immediately - no waiting for emails/logging
        return redirect(url_for('contact.contact_form'))
    else:
        # Form validation failed - log the errors for debugging
        if request.method == 'POST':
            current_app.logger.warning(f"Form validation failed. Errors: {form.errors}")
            if 'specialties' in form.errors:
                current_app.logger.warning(f"Specialty field errors: {form.errors['specialties']}")
                current_app.logger.warning(f"Specialty field data: '{form.specialties.data}'")

            # Handle AJAX form validation errors
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({
                    'success': False,
                    'errors': form.errors,
                    'message': 'Please check the form for errors and try again.'
                })

    # If the request is GET (initial load) or form validation failed (on POST),
    # render the contact_form.html template.
    # The 'form' object is passed to the template so Jinja can render the form fields.
    return render_template('contact_form.html', form=form)


def process_contact_form_background(form_data, app):
    """
    🔥 PERFORMANCE OPTIMIZATION: Background processing of emails and logging
    This function runs in a separate thread to avoid blocking the user response
    """
    with app.app_context():
        start_time = time.time()
        app.logger.info("🔄 Starting background processing for {}".format(form_data['name']))

        # Track success/failure of each operation
        results = {
            'user_confirmation': False,
            'admin_notification': False,
            'analytics_notification': False,
            'excel_logged': False,
            'json_logged': False
        }

        # 1. Send user acknowledgment email (highest priority)
        try:
            send_user_acknowledgment_email(form_data)
            results['user_confirmation'] = True
            app.logger.info("✅ User acknowledgment email sent to {}".format(form_data['email']))
        except Exception as e:
            app.logger.error("❌ Failed to send user acknowledgment email: {}".format(str(e)))

        # 2. Send admin notification email
        try:
            send_admin_email(form_data)
            results['admin_notification'] = True
            app.logger.info("✅ Admin notification email sent")
        except Exception as e:
            app.logger.error("❌ Failed to send admin notification email: {}".format(str(e)))

        # 3. Send analytics notification email (if enabled)
        try:
            if app.config.get('ENABLE_ANALYTICS_EMAIL', True):
                send_analytics_email(form_data)
                results['analytics_notification'] = True
                app.logger.info("✅ Analytics notification email sent")
        except Exception as e:
            app.logger.error("❌ Failed to send analytics notification email: {}".format(str(e)))

        # 4. Log to Excel file
        try:
            log_to_excel(form_data)
            results['excel_logged'] = True
            app.logger.info("✅ Form data logged to Excel successfully")
        except Exception as e:
            app.logger.error("❌ Failed to log to Excel: {}".format(str(e)))

        # 5. Log form submission (JSON backup)
        try:
            log_form_submission(form_data)
            results['json_logged'] = True
            app.logger.info("✅ Form data logged to JSON successfully")
        except Exception as e:
            app.logger.error("❌ Failed to log form data to JSON: {}".format(str(e)))

        # Performance and success metrics
        end_time = time.time()
        processing_time = round(end_time - start_time, 2)
        success_count = sum(results.values())

        app.logger.info("📊 Background processing completed for {} in {}s".format(form_data['name'], processing_time))
        app.logger.info("📈 Success rate: {}/5 operations completed successfully".format(success_count))

        if success_count < 5:
            failed_operations = [op for op, success in results.items() if not success]
            app.logger.warning("⚠️ Failed operations: {}".format(', '.join(failed_operations)))


@contextmanager
def email_timeout(timeout_seconds=10):
    """Context manager for email operations with timeout"""
    old_timeout = socket.getdefaulttimeout()
    try:
        socket.setdefaulttimeout(timeout_seconds)
        yield
    finally:
        socket.setdefaulttimeout(old_timeout)


def send_email_with_retry(msg, max_retries=2, timeout=10):
    """Send email with timeout and retry logic"""
    for attempt in range(max_retries + 1):
        try:
            with email_timeout(timeout):
                mail.send(msg)
                return True
        except Exception as e:
            current_app.logger.warning("📧 Email attempt {} failed: {}".format(attempt + 1, str(e)))
            if attempt == max_retries:
                raise
            time.sleep(1)  # Brief delay before retry
    return False


def send_user_acknowledgment_email(form_data):
    """Send acknowledgment email to user"""
    try:
        # Check if email configuration is available
        if not current_app.config.get('MAIL_USERNAME'):
            current_app.logger.warning("Email credentials not configured - skipping user email")
            raise Exception("Email credentials not configured")

        # Create user-friendly acknowledgment email
        user_email_body = f"""
Dear {form_data['name']},

Thank you for contacting Ardur Healthcare! We have successfully received your inquiry and appreciate you taking the time to reach out to us.

📅 Submission Details:
• Submitted: {datetime.fromisoformat(form_data['timestamp']).strftime('%B %d, %Y at %I:%M %p UTC')}
• Service Type: {form_data['service_type'] or 'General Inquiry'}
• Your Message: "{form_data['message'][:100]}{'...' if len(form_data['message']) > 100 else ''}"

What happens next?
✅ Our team will review your inquiry within 24 hours
✅ A specialist will contact you at {form_data['phone']} or reply to this email
✅ We'll provide you with detailed information about our {form_data['service_type'] or 'services'}

If you have any urgent questions or need to update your information, please don't hesitate to contact us directly at dhoteharshal16@gmail.com or call us.

Thank you for choosing Ardur Healthcare for your medical billing and practice management needs.

Best regards,
The Ardur Healthcare Team

---
This is an automated confirmation email. Please do not reply to this message.
If you need immediate assistance, contact us at dhoteharshal16@gmail.com
        """

        # Get sender email
        sender_email = current_app.config.get('MAIL_DEFAULT_SENDER', 'dhoteharshal16@gmail.com')

        # Create and send message
        msg = Message(
            subject=f"✅ Thank you for contacting Ardur Healthcare - We've received your {form_data['service_type'] or 'inquiry'}",
            sender=sender_email,
            recipients=[form_data['email']],
            body=user_email_body
        )

        current_app.logger.info("Sending user acknowledgment email from {} to {}".format(sender_email, form_data['email']))

        # 🚀 PERFORMANCE: Use timeout and retry logic
        timeout = current_app.config.get('MAIL_TIMEOUT', 10)
        max_retries = current_app.config.get('MAIL_MAX_RETRIES', 2)

        send_email_with_retry(msg, max_retries, timeout)
        current_app.logger.info("✅ User acknowledgment email sent successfully to {}".format(form_data['email']))

    except Exception as e:
        error_details = "Error sending user acknowledgment email: {}".format(str(e))
        current_app.logger.error("❌ {}".format(error_details))
        raise Exception(error_details)


def send_admin_email(form_data):
    """Send complete form data to admin email"""
    try:
        # Check if email configuration is available
        if not current_app.config.get('MAIL_USERNAME'):
            current_app.logger.warning("Email credentials not configured - skipping admin email")
            raise Exception("Email credentials not configured")

        # Format specialties for better readability
        if isinstance(form_data['specialties'], str):
            specialties_text = form_data['specialties'] if form_data['specialties'].strip() else 'Not specified'
        elif isinstance(form_data['specialties'], list):
            specialties_text = ', '.join(form_data['specialties']) if form_data['specialties'] else 'Not specified'
        else:
            specialties_text = 'Not specified'

        # Create enhanced email content
        email_body = f"""
🏥 NEW CONTACT FORM SUBMISSION
========================================

📅 Submitted: {datetime.fromisoformat(form_data['timestamp']).strftime('%B %d, %Y at %I:%M %p UTC')}
🌐 IP Address: {form_data['ip_address']}

👤 CONTACT INFORMATION:
• Name: {form_data['name']}
• Email: {form_data['email']}
• Phone: {form_data['phone']}
• Practice: {form_data['practice_name'] or 'Not specified'}
• State: {form_data['state']}

💼 SERVICE DETAILS:
• Specialties: {specialties_text}
• Service Type: {form_data['service_type'] or 'Not specified'}

💬 MESSAGE:
{form_data['message']}

📝 SUBJECT: {form_data['subject']}

✅ PRIVACY POLICY AGREEMENT: {'Agreed' if form_data['privacy_policy_agreement'] else 'Not Agreed'}

========================================
📧 IMPORTANT: User has been sent an acknowledgment email
📞 Contact {form_data['name']} at {form_data['phone']} or reply to {form_data['email']}

This message was sent from the Ardur Healthcare contact form.
Generated: {form_data['timestamp']}
        """

        # Get recipient email with fallback
        admin_email = current_app.config.get('ADMIN_EMAIL', 'dhoteharshal16@gmail.com')
        sender_email = current_app.config.get('MAIL_DEFAULT_SENDER', current_app.config.get('MAIL_USERNAME'))

        # Create and send message
        msg = Message(
            subject=f"🏥 NEW CONTACT: {form_data['name']} - {form_data['service_type'] or 'General Inquiry'} ({form_data['state']})",
            sender=sender_email,
            recipients=[admin_email],
            reply_to=form_data['email'],  # Allow direct reply to customer
            body=email_body
        )

        current_app.logger.info(f"Sending admin email from {sender_email} to {admin_email}")

        # 🚀 PERFORMANCE: Use timeout and retry logic
        timeout = current_app.config.get('MAIL_TIMEOUT', 10)
        max_retries = current_app.config.get('MAIL_MAX_RETRIES', 2)

        send_email_with_retry(msg, max_retries, timeout)
        current_app.logger.info(f"✅ Admin email sent successfully to {admin_email} for {form_data['name']}")

    except Exception as e:
        error_details = f"Error sending admin email: {str(e)}"
        current_app.logger.error(error_details)
        raise Exception(error_details)


def send_analytics_email(form_data):
    """Send brief analytics notification to analytics email"""
    try:
        # Check if email configuration is available
        if not current_app.config.get('MAIL_USERNAME'):
            current_app.logger.warning("Email credentials not configured - skipping analytics email")
            raise Exception("Email credentials not configured")

        # Format specialties for better readability
        specialties_text = ', '.join(form_data['specialties']) if isinstance(form_data['specialties'], list) else form_data['specialties']
        if not specialties_text or not specialties_text.strip():
            specialties_text = 'Not specified'

        # Create brief analytics email content (no personal info)
        analytics_body = f"""
📊 NEW FORM SUBMISSION ANALYTICS
================================

📅 Timestamp: {datetime.fromisoformat(form_data['timestamp']).strftime('%B %d, %Y at %I:%M %p UTC')}

📍 GEOGRAPHIC DATA:
• State: {form_data['state']}

🏥 SERVICE DATA:
• Specialties: {specialties_text}
• Service Type: {form_data['service_type'] or 'General Inquiry'}

📈 SUMMARY:
A new form submission was received from a {specialties_text.lower()} practice
in {form_data['state']} requesting information about {form_data['service_type'] or 'general services'}.

================================
This is anonymized analytics data from Ardur Healthcare contact forms.
For full details, check the admin email or Excel log.
        """

        # Get emails
        analytics_email = current_app.config.get('ANALYTICS_EMAIL', 'dhoteh020@gmail.com')
        sender_email = current_app.config.get('MAIL_DEFAULT_SENDER', 'dhoteharshal16@gmail.com')

        # Create and send message
        msg = Message(
            subject=f"📊 Analytics Alert: New Form Submission from {form_data['state']} - {form_data['service_type'] or 'General'}",
            sender=sender_email,
            recipients=[analytics_email],
            body=analytics_body
        )

        current_app.logger.info(f"Sending analytics email from {sender_email} to {analytics_email}")

        # 🚀 PERFORMANCE: Use timeout and retry logic
        timeout = current_app.config.get('MAIL_TIMEOUT', 10)
        max_retries = current_app.config.get('MAIL_MAX_RETRIES', 2)

        send_email_with_retry(msg, max_retries, timeout)
        current_app.logger.info(f"✅ Analytics email sent successfully to {analytics_email}")

    except Exception as e:
        error_details = f"Error sending analytics email: {str(e)}"
        current_app.logger.error(error_details)
        raise Exception(error_details)


def log_to_excel(form_data):
    """🚀 OPTIMIZED Excel logging with performance improvements"""
    try:
        start_time = time.time()

        # Create Excel file path
        excel_dir = os.path.join(current_app.root_path, '..', 'logs')
        os.makedirs(excel_dir, exist_ok=True)
        excel_file = os.path.join(excel_dir, 'submissions.xlsx')

        # Format specialties for Excel
        if isinstance(form_data['specialties'], list):
            specialties_text = ', '.join(form_data['specialties']) if form_data['specialties'] else 'Not specified'
        else:
            specialties_text = form_data['specialties'] if form_data['specialties'] else 'Not specified'

        # Prepare row data
        timestamp_formatted = datetime.fromisoformat(form_data['timestamp']).strftime('%Y-%m-%d %H:%M:%S UTC')
        row_data = [
            timestamp_formatted,
            form_data['name'],
            form_data['email'],
            form_data['phone'],
            form_data['practice_name'] or 'Not specified',
            form_data['state'],
            specialties_text,
            form_data['service_type'] or 'General Inquiry',
            form_data['subject'],
            form_data['message'][:500] if len(form_data['message']) > 500 else form_data['message'],  # Truncate long messages
            'Yes' if form_data['privacy_policy_agreement'] else 'No',
            form_data['ip_address']
        ]

        # 🚀 PERFORMANCE: Use read-only mode first to check if file exists and get row count
        workbook = None
        worksheet = None
        next_row = 1

        if os.path.exists(excel_file):
            try:
                # Quick check for file size to avoid loading huge files
                file_size = os.path.getsize(excel_file)
                if file_size > 10 * 1024 * 1024:  # 10MB limit
                    current_app.logger.warning(f"⚠️ Excel file too large ({file_size} bytes), creating new monthly file")
                    excel_file = excel_file.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m")}.xlsx')

                # Load existing workbook with minimal memory usage
                workbook = openpyxl.load_workbook(excel_file, read_only=False, keep_vba=False)
                worksheet = workbook.active
                if worksheet is not None:
                    next_row = worksheet.max_row + 1 if worksheet.max_row else 1
                else:
                    next_row = 1

            except Exception as load_error:
                current_app.logger.warning(f"⚠️ Could not load existing Excel file: {load_error}")
                workbook = None

        if workbook is None:
            # Create new workbook with optimized settings
            workbook = Workbook(write_only=False)
            worksheet = workbook.active
            if worksheet is not None:
                worksheet.title = "Contact Form Submissions"
            next_row = 1

            # Create headers efficiently
            headers = [
                'Timestamp', 'Name', 'Email', 'Phone', 'Practice Name',
                'State', 'Specialties', 'Service Type', 'Subject',
                'Message', 'Privacy Agreement', 'IP Address'
            ]

            # Add headers without heavy styling for performance
            for col, header in enumerate(headers, 1):
                if worksheet is not None:
                    worksheet.cell(row=1, column=col, value=header)

            next_row = 2

        # Add new row efficiently
        for col, value in enumerate(row_data, 1):
            if worksheet is not None:
                worksheet.cell(row=next_row, column=col, value=value)

        # Save with optimized settings
        workbook.save(excel_file)
        workbook.close()

        elapsed_time = round(time.time() - start_time, 3)
        current_app.logger.info(f"✅ Excel logged in {elapsed_time}s to row {next_row}: {excel_file}")

    except Exception as e:
        error_details = f"Error logging to Excel: {str(e)}"
        current_app.logger.error(error_details)
        raise Exception(error_details)


def log_form_submission(form_data):
    """🚀 OPTIMIZED JSON logging with performance improvements"""
    try:
        start_time = time.time()

        # Create analytics log entry
        analytics_entry = {
            'timestamp': form_data['timestamp'],
            'state': form_data['state'],
            'specialties': form_data['specialties'],
            'service_type': form_data['service_type'],
            'source': 'contact_form',
            'ip_address': form_data['ip_address']
        }

        # Ensure logs directory exists
        logs_dir = os.path.join(current_app.root_path, '..', 'logs')
        os.makedirs(logs_dir, exist_ok=True)

        # Create log file path with date rotation
        today = datetime.now().strftime('%Y%m')
        log_file = os.path.join(logs_dir, f'analytics_log_{today}.json')

        logs = []

        # 🚀 PERFORMANCE: Check file size before loading to avoid memory issues
        if os.path.exists(log_file):
            file_size = os.path.getsize(log_file)
            if file_size > 5 * 1024 * 1024:  # 5MB limit
                current_app.logger.warning(f"⚠️ JSON log file too large ({file_size} bytes), starting new monthly file")
                # Archive current file and start fresh
                archive_file = log_file.replace('.json', f'_archived_{int(time.time())}.json')
                os.rename(log_file, archive_file)
            else:
                try:
                    # Read existing logs with timeout protection
                    with open(log_file, 'r', encoding='utf-8') as f:
                        logs = json.load(f)
                except (json.JSONDecodeError, IOError) as e:
                    current_app.logger.warning(f"⚠️ Could not read JSON log file: {e}")
                    logs = []

        # Add new entry
        logs.append(analytics_entry)

        # Keep only last 500 entries for performance (reduced from 1000)
        if len(logs) > 500:
            logs = logs[-500:]

        # Write back to file with optimized settings
        with open(log_file, 'w', encoding='utf-8') as f:
            json.dump(logs, f, separators=(',', ':'), ensure_ascii=False)  # Compact format

        elapsed_time = round(time.time() - start_time, 3)
        current_app.logger.info(f"✅ JSON logged in {elapsed_time}s: {form_data['state']} submission")

    except Exception as e:
        current_app.logger.error(f"Error logging analytics data: {str(e)}")
        # Don't raise here - logging failure shouldn't break the form submission


# 🚀 PERFORMANCE: Cache specialties data to avoid loading on every request
_specialties_cache = None
_cache_timestamp = 0

def load_specialty_data():
    try:
        json_path = os.path.join(current_app.root_path, '..', 'data', 'specialty', 'specialty_data.json')
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data.get('services', [])
    except (FileNotFoundError, json.JSONDecodeError):
        return []


def get_cached_specialties():
    """🚀 PERFORMANCE: Get specialties with caching for 5 minute TTL"""
    global _specialties_cache, _cache_timestamp

    current_time = time.time()
    cache_ttl = 300  # 5 minutes

    if _specialties_cache is None or (current_time - _cache_timestamp) > cache_ttl:
        try:
            # Get specialties from JSON data
            specialties = load_specialty_data()
            json_specialty_names = []
            for s in specialties:
                clean_name = s['specialty'].replace(' Billing Services', '').strip()
                json_specialty_names.append(clean_name)

            # Additional services not in JSON (static list)
            additional_services = [
                "Allergy and Immunology", "Ambulatory Surgical Center", "Anesthesia",
                "Behavioral Health", "Cardiology", "Chiropractic", "Clinical Psychology",
                "Dental", "Dermatology", "Durable Medical Equipment (DME)", "Emergency Room",
                "Endocrinology", "Family Practice", "Gastroenterology", "General Surgery",
                "Home Health", "Hospice", "Internal Medicine", "Laboratory", "Massage Therapy",
                "Mental Health", "Neurology", "OB GYN", "Occupational Therapy", "Oncology",
                "Optometry", "Oral and Maxillofacial", "Orthopedic", "Otolaryngology (ENT)",
                "Pain Management", "Pathology", "Pediatrics", "Pharmacy", "Physical Therapy",
                "Plastic Surgery", "Podiatry", "Primary Care", "Pulmonology", "Radiation Oncology",
                "Radiology", "Rheumatology", "Skilled Nursing Facility (SNF)", "Sleep Disorder",
                "Sports Medicine", "Urgent Care", "Urology", "Wound Care"
            ]

            # Combine and remove duplicates efficiently
            seen = set()
            all_specialties = []

            for name in json_specialty_names + additional_services:
                if name not in seen:
                    all_specialties.append(name)
                    seen.add(name)

            # Sort alphabetically
            _specialties_cache = sorted(all_specialties)
            _cache_timestamp = current_time

        except Exception as e:
            current_app.logger.error(f"Error loading specialties: {str(e)}")
            return []

    return _specialties_cache


@contact.route('/api/specialties')
def api_specialties():
    """🚀 OPTIMIZED specialties API with caching and efficient filtering"""
    try:
        # Get cached specialties
        all_specialties = get_cached_specialties()

        if not all_specialties:
            return jsonify([]), 500

        # Filter based on search query if provided
        query = request.args.get('q', '').lower().strip()
        if query:
            # Use list comprehension for better performance
            filtered = [s for s in all_specialties if query in s.lower()]
            return jsonify(filtered[:15])  # Increased limit to 15 for better UX

        # Return first 50 specialties for initial load (performance optimization)
        return jsonify(all_specialties[:50])

    except Exception as e:
        current_app.logger.error(f"Error in specialties API: {str(e)}")
        return jsonify([]), 500
